import json

import xlsxwriter
from lxml import etree
from openpyxl import load_workbook

try:
    import cPickle as pickle
except ImportError:
    import pickle


def dumpObj(path, obj):
    with open(path, 'wb') as f:
        pickle.dump(obj, f)


def load(path):
    with open(path, 'rb') as f:
        return pickle.load(f)


def writeXlsx(names, lines, xlsx_file):
    """
    Write a list of lines to an xlsx file.
    """
    workbook = xlsxwriter.Workbook(xlsx_file)
    worksheet = workbook.add_worksheet()
    worksheet.write_row(0, 0, names)
    rownum = 1
    for line in lines:
        worksheet.write_row(rownum, 0, line)
        rownum += 1
    workbook.close()


def readXlsx(xlsx_file):
    """
    Read lines from a xlsx file.
    """
    lines = []
    workbook = load_workbook(xlsx_file, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    for i, row in enumerate(worksheet.rows):
        # if i >= 1:
        line = []
        for col in row:
            s = '' if col.value is None else str(col.value).strip()
            line.append(s)
        lines.append(line)
        # line = [ str(col.value).strip() for col in row ]
        # lines.append(line)
    workbook.close()
    return lines


def writeJson(obj, json_file):
    """
    Write an obj to a Json file.
    """
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(obj, f, indent=4, ensure_ascii=False, sort_keys=True)


def readJson(json_file: object):
    """
    Read an obj from a Json file.
    """
    with open(json_file, 'r', encoding='utf-8') as f:
        obj = json.load(f)
    return obj


def writeDictList2xml(xml_file, dicts, rootname):
    """
    Write a list of dicts (an info unit) to an XML file.
    """
    root = etree.Element(rootname)
    # root = etree.Element(item_name, attrib={'Count': str(len(dict_list))})
    for _dict in dicts:
        try:
            etree.SubElement(root, 'row', attrib=_dict)
        except Exception as e:
            print('***** ERROR in writeDictList2xml():', e, '->', _dict)
    tree = etree.ElementTree(root)
    tree.write(xml_file, pretty_print=True, xml_declaration=True, encoding='utf-8')


def readTxt(txt_file):
    """
    Read the lines of a txt.
    """
    lines = []
    with open(txt_file, 'r', encoding='utf-8') as f:
        for line in f.readlines():
            lines.append(line.strip())
    return lines


def generatePercentageStr(per, acc):
    """
    Generate a percentage str with a specified accuracy.
    """
    if acc == '':
        acc = 2
    acc = '%.' + str(acc) + 'f'
    return  acc % (100 * per) + '%'
